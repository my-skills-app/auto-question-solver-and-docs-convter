"""Export solved questions to Excel (.xlsx) matching CSV template columns."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from .csv_template import CSV_HEADERS, questions_to_rows


def save_xlsx(
    questions: list[dict[str, Any]],
    path: Path,
    *,
    set_name: str = "Paper Name",
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = questions_to_rows(questions, set_name=set_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Solved Questions"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    wrap = Alignment(wrap_text=True, vertical="top")

    for col, name in enumerate(CSV_HEADERS, start=1):
        cell = ws.cell(1, col, name)
        cell.font = header_font
        cell.fill = header_fill

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, key in enumerate(CSV_HEADERS, start=1):
            cell = ws.cell(r_idx, c_idx, row.get(key, ""))
            cell.alignment = wrap

    widths = {
        "question_r": 10,
        "question_hi": 40,
        "question_en": 40,
        "solution_hi": 36,
        "solution_en": 36,
        "answer": 14,
        "set_name": 14,
        "difficulty_level": 12,
    }
    for col, name in enumerate(CSV_HEADERS, start=1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = widths.get(
            name, 18
        )

    ws.freeze_panes = "A2"
    try:
        wb.save(path)
        return path
    except PermissionError:
        alt = path.with_name(f"{path.stem}.checkpoint{path.suffix}")
        wb.save(alt)
        return alt
